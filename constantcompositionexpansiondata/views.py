from django.http import JsonResponse
from django.shortcuts import render, redirect
from IntelligentOilWell.custom_context_processors import selectedfgi
from django.contrib import messages
from .models import  CCEPVTData
from .forms import CCEPVTDATAForm
from tablib import Dataset
from constantcompositionexpansion.resources import CCEPVTdataResource
from constantcompositionexpansion.models import CCEPVT
from openpyxl import load_workbook
from .utils import get_Multiplot
import numpy as np

# Create your views here.
def list_ccepvtdata(request, ccepvt):   
    Pb=0
    ccepvtdatas = CCEPVTData.objects.filter(ccepvt = ccepvt).all()   
    model = CCEPVT.objects.get(pk=ccepvt)
    temp = model.temperature
    x=[]
    y=[] 
    x1=[]
    y1=[] 
    #x1=[x.pressure for x in ccepvtdatas]
    #y1=[y.relative_volume  for y in ccepvtdatas]
    #Pb= [x.pressure for x in ccepvtdatas ]
    for data in ccepvtdatas:
        if data.y_function !=None:
            x.append(data.pressure)
            y.append(data.y_function) 
            
    for data in ccepvtdatas:
        if data.relative_volume <=1:
            x1.append(data.pressure)
            y1.append(data.relative_volume) 
    for data in ccepvtdatas:
        if data.relative_volume ==1:
            Pb = data.pressure 
            density_Pb =data.density
            break
    m, b = np.polyfit(x, y, 1)
    m= round(m,5)
    b= round(b,5)
    print  (f"equation : y = {m}x  + {b}")  
    print  (Pb, density_Pb, temp)  
    co = calculate_co(2000, 2500,ccepvtdatas )
    print(co)
    chart = get_Multiplot(x,y, x1, y1, Pb)   
    return render (request, 'constantcompositionexpansiondata/ccePVTdata.html', {'ccepvt':ccepvt, 'ccepvtdatas': ccepvtdatas, 'chart':chart})
    
def create_ccepvtdata(request,ccepvt): 
   ccepvt = CCEPVT.objects.get(id=ccepvt)
   ccepvtdata = CCEPVTData()
   ccepvtdata.ccepvt = ccepvt
   form = CCEPVTDATAForm(request.POST or None, instance=ccepvtdata)  
   if form.is_valid():
       form.save()       
       return redirect ('constantcompositionexpansiondata:list_ccepvtdata', ccepvt )    
   return render (request, 'constantcompositionexpansiondata/ccePVTdata_form.html', {'form': form, 'ccepvt':ccepvt})

def update_ccepvtdata(request, id):
   ccepvtdata = CCEPVTData.objects.get(id=id)
   ccepvt =(ccepvtdata.ccepvt).pk 
   form = CCEPVTDATAForm(request.POST or None, instance=ccepvtdata)
   if form.is_valid():
       form.save()
       return redirect ('constantcompositionexpansiondata:list_ccepvtdata', ccepvt)
   return render (request, 'constantcompositionexpansiondata/ccePVTdata_form.html', {'form': form, 'ccepvtdata':ccepvtdata, 'id':id})

def delete_ccepvtdata(request, id):
   ccepvtdata = CCEPVTData.objects.get(id=id)  
   ccepvt =(ccepvtdata.ccepvt).pk  
   if request.method == 'POST' :
       ccepvtdata.delete()
       return redirect ('constantcompositionexpansiondata:list_ccepvtdata',ccepvt )
   return render (request, 'constantcompositionexpansiondata/ccePVTdata_confirm_delete.html', {'ccepvtdata':ccepvtdata, 'id':id})

def excel_upload_cceData(request,ccepvt ):    
    print("I am in")
    if request.method=="POST":
        ccepvtdataresource =CCEPVTdataResource()        
        dataset =Dataset()
        ccepvt =ccepvt
        new_ccedata = request.FILES['myfile']
        if not new_ccedata.name.endswith('xlsx'):
            messages.info(request,'wrong format')
            return render(request, 'constantcompositionexpansiondata/ccePVTdata.html', ccepvt)
        imported_data = dataset.load(new_ccedata.read(), format='xlsx')
        print(imported_data)
        for data in imported_data:
            value = new_ccedata(
                ccepvt,
                data[0],
                data[1],
                data[2],
                data[3]
                         )
            value.save()
        return render(request, 'constantcompositionexpansiondata/ccePVTdata.html', ccepvt)

def upload(request, ccepvt):    
    ccepvtdata = CCEPVT.objects.get(id=ccepvt) 
    ccpvtdatas = CCEPVTData.objects.filter(ccepvt=ccepvtdata)
    ccpvtdatas.delete()
    print("I am in Get")
    if request.method == 'POST' and request.FILES['myfile']:
        print("I am in Post")
        myfile = request.FILES['myfile']
        wb = load_workbook(filename=myfile, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            obj = CCEPVTData()
            obj.ccepvt = ccepvtdata
            obj.pressure =row[0].value 
            obj.relative_volume=row[1].value 
            obj.y_function=row[2].value 
            obj.density=row[3].value
            obj.save()
        return redirect('constantcompositionexpansiondata:list_ccepvtdata', ccepvt )  
    return render(request, 'constantcompositionexpansiondata/ccePVTdata.html', {"ccepvt":ccepvt})


def calculate_co(P1, P2, datas):
    v1=0
    v2=0
    p1=0
    p2=0
    for i  in range(len(datas)):        
        if i >1 and P1 >= datas[i].pressure:
            p1 = datas[i-1].pressure
            v1 = datas[i-1].relative_volume           
            break
    for i  in range(len(datas)):        
        if i >1 and P2 >= datas[i].pressure:          
            p2 = datas[i].pressure
            v2 = datas[i].relative_volume
            break
    co = (-1/v2) *(v2-v1)/(p2-p1)        
    return co
      